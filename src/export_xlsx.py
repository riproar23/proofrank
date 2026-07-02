"""
Convert a validated ProofRank ranked CSV to .xlsx.

Usage
-----
    python -m src.export_xlsx                          # default paths
    python -m src.export_xlsx output/Viburnia.csv output/Viburnia.xlsx
    python -m src.export_xlsx some_other.csv out.xlsx

Rules
-----
- Candidate IDs (CAND_...) stay as TEXT — never coerced to numbers.
- A cell becomes int   if str(int(s))   == s  (exact round-trip).
- A cell becomes float if str(float(s)) == s  (exact round-trip).
- Everything else stays as a plain string.
- Header row is bold and frozen at A2.
- Columns are auto-sized, capped at 48 characters wide.
"""

import csv
import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).parent.parent
DEFAULT_SRC = ROOT / "output" / "Viburnia.csv"
DEFAULT_DST = ROOT / "output" / "Viburnia.xlsx"

# Columns whose values must always stay as strings regardless of content.
# Add column names here if the CSV ever grows new ID-like columns.
_FORCE_TEXT = {"candidate_id"}

COL_MAX_WIDTH = 48


def _coerce(value: str, col_name: str):
    """Return value coerced to int/float only when the round-trip is exact."""
    if col_name in _FORCE_TEXT or not value:
        return value
    try:
        as_int = int(value)
        if str(as_int) == value:
            return as_int
    except ValueError:
        pass
    try:
        as_float = float(value)
        if str(as_float) == value:
            return as_float
    except ValueError:
        pass
    return value


def convert(src: Path, dst: Path) -> None:
    if not src.exists():
        sys.exit(f"ERROR: input file not found: {src}")

    with src.open(newline="", encoding="utf-8") as fh:
        reader = csv.reader(fh)
        try:
            headers = next(reader)
        except StopIteration:
            sys.exit(f"ERROR: input file is empty: {src}")
        rows = list(reader)

    if not rows:
        sys.exit(f"ERROR: input file has a header but no data rows: {src}")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Shortlist"

    bold = Font(bold=True)

    # Write header
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = bold

    # Write data rows
    for row_idx, row in enumerate(rows, start=2):
        for col_idx, (header, raw) in enumerate(zip(headers, row), start=1):
            ws.cell(row=row_idx, column=col_idx, value=_coerce(raw, header))

    # Freeze header
    ws.freeze_panes = "A2"

    # Auto-size columns
    for col_idx, header in enumerate(headers, start=1):
        col_letter = get_column_letter(col_idx)
        # Measure widest content in this column (header + all data cells)
        col_values = [header] + [r[col_idx - 1] for r in rows if col_idx - 1 < len(r)]
        max_len = max((len(str(v)) for v in col_values), default=10)
        ws.column_dimensions[col_letter].width = min(max_len + 2, COL_MAX_WIDTH)

    dst.parent.mkdir(parents=True, exist_ok=True)
    wb.save(dst)

    print(
        f"Wrote {len(rows)} data rows x {len(headers)} columns -> {dst}"
    )


def main() -> None:
    args = sys.argv[1:]
    src = Path(args[0]) if len(args) >= 1 else DEFAULT_SRC
    dst = Path(args[1]) if len(args) >= 2 else DEFAULT_DST
    convert(src, dst)


if __name__ == "__main__":
    main()
